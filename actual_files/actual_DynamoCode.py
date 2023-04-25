import sys
import os
import clr
import sys
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

localapp = os.getenv(r'LOCALAPPDATA')
sys.path.append(os.path.join(localapp, r'python-3.8.3-embed-amd64\Lib\site-packages'))

import System
from System import Array
from System.Collections.Generic import *

clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *

clr.AddReference("RevitNodes")
import Revit

clr.ImportExtensions(Revit.Elements)
clr.ImportExtensions(Revit.GeometryConversion)

clr.AddReference("RevitServices")
import RevitServices
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

clr.AddReference("RevitAPI")
clr.AddReference("RevitAPIUI")

import Autodesk
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *

doc = DocumentManager.Instance.CurrentDBDocument
uiapp = DocumentManager.Instance.CurrentUIApplication
app = uiapp.Application
uidoc = uiapp.ActiveUIDocument

floors_up = {}
floors_down = {}
precast_elements = {}
columns = {}
beams = {}
slurry_Bisus = {}
slurry_Dipun = {}
walls_in_new = {}
walls_out_new = {}
Dipuns = {}
Bisus = {}
Basic_Plate = {}
Found_Head = {}
Rafsody = {}
Slabedge = {}
Found_without_DTM = {}
"""________Stairs________"""

stairs_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_Stairs). \
    WhereElementIsNotElementType(). \
    ToElements()

floors_collector_forsteirs = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_Floors). \
    WhereElementIsNotElementType(). \
    ToElements()

Stairs = {}


def getiing_parameters_slab_edge(stairs_collector, floors_collector_forsteirs):
    for el in stairs_collector:
        stair_type_id = el.GetTypeId()
        stair_type_elem = doc.GetElement(stair_type_id)
        parameter_Duplication = stair_type_elem.LookupParameter("Duplication Type Mark").AsString()
        if stair_type_elem:
            if parameter_Duplication == "Home" or parameter_Duplication == "Stairway" or parameter_Duplication == "Stairs":
                parameter_vol = el.LookupParameter("Volume")
                key = "Stairs/משטחי מדרגות ישרים (פודסטי ביניים ) ומדרגות מבטון"
                if key not in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key] = {"Volume": parameter_value_vol}
                elif key in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key]["Volume"] += parameter_value_vol
            elif parameter_Duplication == "Demolished":
                parameter_vol = el.LookupParameter("Volume")
                key = "Stairs-Demolished/מדרגות-הרוס"
                if key not in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key] = {"Volume": parameter_value_vol}
                elif key in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key]["Volume"] += parameter_value_vol
            else:
                pass

    for el in floors_collector_forsteirs:
        floor_type_id = el.GetTypeId()
        floor_type_elem = doc.GetElement(floor_type_id)
        parameter_Duplication = floor_type_elem.LookupParameter("Duplication Type Mark").AsString()
        if floor_type_elem:
            if parameter_Duplication == "Landing-H" or parameter_Duplication == "Landing-S":
                parameter_vol = el.LookupParameter("Volume")
                key = "Stairs/משטחי מדרגות ישרים (פודסטי ביניים ) ומדרגות מבטון"
                if key not in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key] = {"Volume": parameter_value_vol}
                elif key in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key]["Volume"] += parameter_value_vol
            if parameter_Duplication == "Demolished":
                parameter_vol = el.LookupParameter("Volume")
                key = "Stairs-Demolished/מדרגות-הרוס"
                if key not in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key] = {"Volume": parameter_value_vol}
                elif key in Stairs:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Stairs[key]["Volume"] += parameter_value_vol


getiing_parameters_slab_edge(stairs_collector, floors_collector_forsteirs)

"""________SlabEdge________"""
slab_edge_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_EdgeSlab). \
    WhereElementIsNotElementType(). \
    ToElements()

def getiing_parameters_slab_edge(slab_edge_collector):
    for el in slab_edge_collector:
        edge_type_id = el.GetTypeId()
        edge_type_elem = doc.GetElement(edge_type_id)
        parameter_Duplication = edge_type_elem.LookupParameter("Duplication Type Mark").AsString()
        if edge_type_elem:
            if not parameter_Duplication:
                parameter_vol = el.LookupParameter("Volume")
                parameter_length = el.LookupParameter("Length")
                key_slab_edge = "no DTM Slab edge"
                if key_slab_edge not in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key_slab_edge] = {"Volume": parameter_value_vol, "Length": parameter_length}
                elif key_slab_edge in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key_slab_edge]["Volume"] += parameter_value_vol
                    Slabedge[key_slab_edge]["Length"] += parameter_length
            elif parameter_Duplication == "Wuta":
                parameter_vol = el.LookupParameter("Volume")
                parameter_length = el.LookupParameter("Length")
                key = "Wuta/ווטות מתחת לרצפה"
                if key not in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key] = {"Volume": parameter_value_vol, "Length": parameter_length}
                elif key in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key]["Volume"] += parameter_value_vol
                    Slabedge[key]["Length"] += parameter_length

            elif parameter_Duplication == "CLSM":
                parameter_vol = el.LookupParameter("Volume")
                parameter_length = el.LookupParameter("Length")
                key = "CLSM מתחת לרצפה"
                if key not in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key] = {"Volume": parameter_value_vol, "Length": parameter_length}
                elif key in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[key]["Volume"] += parameter_value_vol
                    Slabedge[key]["Length"] += parameter_length

            else:
                parameter_vol = el.LookupParameter("Volume")
                parameter_length = el.LookupParameter("Length")
                if parameter_Duplication not in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[parameter_Duplication] = {"Volume": parameter_value_vol, "Length": parameter_length}
                elif parameter_Duplication in Slabedge:
                    parameter_length = parameter_length.AsDouble() * 0.3048
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    Slabedge[parameter_Duplication]["Volume"] += parameter_value_vol
                    Slabedge[parameter_Duplication]["Length"] += parameter_length


getiing_parameters_slab_edge(slab_edge_collector)

"""________COLUMNS________"""
columns_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_StructuralColumns). \
    WhereElementIsNotElementType(). \
    ToElements()

def getiing_parameters(columns_collector):
    for el in columns_collector:
        col_type_id = el.GetTypeId()
        col_type_elem = doc.GetElement(col_type_id)
        if col_type_elem:
            parameter_Duplication = col_type_elem.LookupParameter("Duplication Type Mark").AsString()
            if parameter_Duplication == "Steel":
                continue

            if parameter_Duplication in ["Rec", "Round", "Eliptic"]:
                parameter_vol = el.LookupParameter("Volume")
                key = "Columns Regular/עמודי בטון"
                if key not in columns:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    columns[key] = {"Volume": parameter_value_vol}
                elif key in columns:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    columns[key]["Volume"] += parameter_value_vol

            if parameter_Duplication == "Precast":
                key_precast = "עמוד טרומי/Columns Precast"
                parameter_vol = el.LookupParameter("Volume")
                if key_precast not in precast_elements:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    precast_elements[key_precast] = {"Volume": parameter_value_vol, "Count": 1}
                elif key_precast in precast_elements:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    precast_elements[key_precast]["Volume"] += parameter_value_vol
                    precast_elements[key_precast]["Count"] += 1

            if not parameter_Duplication:
                key = "DTM empty Columns"
                parameter_vol = el.LookupParameter("Volume")
                if key not in columns:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    columns[key] = {"Volume": parameter_value_vol}
                else:
                    parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                    columns[key]["Volume"] += parameter_value_vol



getiing_parameters(columns_collector)
"""________FLOORS________"""
floors_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_Floors). \
    WhereElementIsNotElementType(). \
    ToElementIds()


def getting_floors_parameters(floor_list):
    for el in floor_list:
        floor_element = doc.GetElement(el)
        floor_type = floor_element.FloorType
        floor_type_comments = floor_type.get_Parameter(BuiltInParameter.ALL_MODEL_TYPE_COMMENTS).AsString()
        floor_duplicationTypeMark = floor_type.LookupParameter("Duplication Type Mark").AsString()
        if floor_type_comments == "Up":
            if not floor_duplicationTypeMark:
                key = "DTM empty Up_floors"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark in ["Total Floor Area", "Total Floor Area Commercial",
                                               "Total Floor Area LSP",
                                               "Total Floor Area Pergola",
                                               "Air Double Level", "Air Elevator", "Air Pergola Aluminium",
                                               "Air Pergola Steel", "Air Pergola Wood", "Air Regular", "Air Stairs",
                                               "Landing-H", "Landing-S", "Landing Steel", "Polivid", "Backfilling",
                                               "Aggregate"]:
                continue
            elif floor_duplicationTypeMark in ["Regular", "Balcon","Terasa"]:
                combined_key = "Regular_up/תקרת בטון"
                if combined_key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[combined_key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[combined_key]['Area'] += floor_area
                    floors_up[combined_key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Regular-T":
                combined_key = "Transformation/תקרת טרנספורמציה מבטון"
                if combined_key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[combined_key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[combined_key]['Area'] += floor_area
                    floors_up[combined_key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Regular-P":
                key = "תקרה דרוכה/Prestressed slab"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Regular-W Special":
                key = "תקרת צלעות/Ribbed slab"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Koteret":
                key = "כותרת לעיבוי נגד חדירה/Koteret"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Rampa":
                key = "רמפה/Rampa"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Geoplast-D":
                key = "תקרת צלעות במילוי גיאופלסט/Geoplast"
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume
            elif floor_duplicationTypeMark == "Slab":
                key = 'לוח"ד/Hollow slabs'
                if key not in precast_elements:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    precast_elements[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    precast_elements[key]['Area'] += floor_area
                    precast_elements[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Completion":
                key = 'השלמות יציקה בין לוחדים/Complitions_up'
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume
            elif floor_duplicationTypeMark == "Topping":
                key = 'יציקת טופינג/Topping'
                if key not in floors_up:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_up[key]['Area'] += floor_area
                    floors_up[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark not in floors_up:
                floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                floors_up[floor_duplicationTypeMark] = {"Area": floor_area, "Volume": floor_volume}

            else:
                floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                floors_up[floor_duplicationTypeMark]["Area"] += floor_area
                floors_up[floor_duplicationTypeMark]["Volume"] += floor_volume

        elif floor_type_comments == "Down":
            if not floor_duplicationTypeMark:
                key = "DTM empty DN_floors"
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark in ["Total Floor Area", "Total Floor Area Commercial",
                                               "Total Floor Area LSP",
                                               "Total Floor Area Pergola",
                                               "Air Double Level", "Air Elevator", "Air Pergola Aluminium",
                                               "Air Pergola Steel", "Air Pergola Wood", "Air Regular", "Air Stairs",
                                               "Landing-H", "Landing-S", "Landing Steel", "Polivid", "Backfilling",
                                               "Aggregate"]:
                continue

            elif floor_duplicationTypeMark in ["Regular", "Balcon", "Terasa"]:
                combined_key = "Regular_dn/רצפת בטון"
                if combined_key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[combined_key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[combined_key]['Area'] += floor_area
                    floors_down[combined_key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Rampa":
                key = 'רמפה/Rampa_dn'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Slab-Rib Special":
                key = 'רצפת צלעות/Ribbed floor'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Supporting":
                key = 'רגל לקיר תומך/leg for supporting wall'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Lite Beton":
                key = 'מצע בטון רזה מתחת לרצפת בטון/Lite Beton'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Completion":
                key = 'השלמות יציקה מעל קורות/Dn_Complition'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "Sidewalk":
                key = 'מדרכות/Sidewalks'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark == "CLSM":
                key = 'CLSM'
                if key not in floors_down:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key] = {"Area": floor_area, "Volume": floor_volume}
                else:
                    floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                    floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                    floors_down[key]['Area'] += floor_area
                    floors_down[key]['Volume'] += floor_volume

            elif floor_duplicationTypeMark not in floors_down:
                floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                floors_down[floor_duplicationTypeMark] = {"Area": floor_area, "Volume": floor_volume}

            else:
                floor_area = floor_element.LookupParameter("Area").AsDouble() * 0.092903
                floor_volume = floor_element.LookupParameter("Volume").AsDouble() * 0.0283168466
                floors_down[floor_duplicationTypeMark]["Area"] += floor_area
                floors_down[floor_duplicationTypeMark]["Volume"] += floor_volume

    return floors_up, floors_down


getting_floors_parameters(floors_collector)

"""________BEAMS________"""
beams_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_StructuralFraming). \
    WhereElementIsNotElementType(). \
    ToElements()


def beams_parameters(beam_list):
    for el in beam_list:
        element_type = doc.GetElement(el.GetTypeId())
        custom_param = element_type.LookupParameter("Duplication Type Mark").AsString()

        if custom_param == "Beam Steel":
            pass
        elif custom_param == "Beam Anchor":
            key = "Beam Anchor/קורת עוגנים"
            if key not in beams:
                beams[key] = {"Count": 1}
            else:
                beams[key]['Count'] += 1

        elif custom_param == "Anchor Polymer":
            key = "עוגנים פולימרים/Anchor Polymer"
            if key not in beams:
                beams[key] = {"Count": 1}
            else:
                beams[key]['Count'] += 1

        elif custom_param == "Anchor Steel":
            key = "עוגנים/Anchor Steel"
            if key not in beams:
                beams[key] = {"Count": 1}
            else:
                beams[key]['Count'] += 1

        elif custom_param == "Precast":
            key = "קורה טרומית/Precast"
            if key not in precast_elements:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                precast_elements[key] = {"Volume": beam_volume, "Count": 1}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                precast_elements[key]['Volume'] += beam_volume
                precast_elements[key]['Count'] += 1

        elif custom_param == "Concrete":
            key = "קורת עוגנים מבטון/Beam Anchor Concrete "
            if key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key] = {"Volume": beam_volume, "Count": 1}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key]['Volume'] += beam_volume
                beams[key]['Count'] += 1

        elif custom_param == "Foundation":
            key = "קורת יסוד/Foundation "
            if key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key] = {"Volume": beam_volume}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key]['Volume'] += beam_volume

        elif custom_param == "Head":
            key = "קורת ראש כלונסאות דיפו/ Foundation Head "
            if key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key] = {"Volume": beam_volume}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key]['Volume'] += beam_volume

        elif custom_param == "Prestressed":
            key = "קורת דרוכה/Beam presstressed"
            if key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key] = {"Volume": beam_volume}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key]['Volume'] += beam_volume

        elif custom_param in ["Regular", "Balcon", "Transformation", "Pergola", "Belts", "Tapered","Regular-T"]:
            combined_key = "קורות בטון/Regular beams"
            if combined_key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[combined_key] = {"Volume": beam_volume}
            elif combined_key in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[combined_key]['Volume'] += beam_volume
            # else:
            #     beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
            #     beams[combined_key] = {"Volume": beam_volume}
        elif not custom_param:
            key = "Without Duplication Type Mark"
            if key not in beams:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key] = {"Volume": beam_volume}
            else:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[key]["Volume"] += beam_volume

        elif custom_param not in beams:
            # key = "קורות/Beams"
            beam_volume = el.LookupParameter("Volume").AsDouble()
            if beam_volume:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[custom_param] = {"Volume": beam_volume, "Count": 1}
        else:
            try:
                beam_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                beams[custom_param]["Volume"] += beam_volume
                beams[custom_param]["Count"] += 1
                # beams[custom_param]["Count"] += 1
            except:
                pass

    return beams


beams_parameters(beams_collector)

"""________Walls________"""
wall_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_Walls). \
    WhereElementIsNotElementType(). \
    ToElementIds()


def getting_Area_Volume_walls(walls_list):
    for wall in wall_collector:
        new_w = doc.GetElement(wall)
        # going to WallType
        wall_type = new_w.WallType
        volume_param = new_w.LookupParameter("Volume")
        area_param = new_w.LookupParameter("Area")
        wall_duplicationTypeMark = wall_type.LookupParameter("Duplication Type Mark").AsString()
        wall_type_comments = wall_type.get_Parameter(BuiltInParameter.ALL_MODEL_TYPE_COMMENTS).AsString()

        if wall_duplicationTypeMark == "Slurry Bisus":
            key = "Slurry Bisus/קיר סלארי ביסוס"
            if key in slurry_Bisus:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                slurry_Bisus[key]["Area"] += wall_area
                slurry_Bisus[key]["Volume"] += wall_volume
            elif key not in slurry_Bisus:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                slurry_Bisus[key] = {"Area": wall_area, "Volume": wall_volume}

        if wall_duplicationTypeMark == "Slurry Dipun":
            key = "Slurry Dipun/קיר סלארי דיפון"
            if key in slurry_Dipun:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                slurry_Dipun[key]["Area"] += wall_area
                slurry_Dipun[key]["Volume"] += wall_volume
            elif key not in slurry_Dipun:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                slurry_Dipun[key] = {"Area": wall_area, "Volume": wall_volume}

        if wall_duplicationTypeMark == "Demolished":
            key = "Demolished/נהרס"
            if key not in walls_in_new:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                walls_in_new[key] = {"Area": wall_area, "Volume": wall_volume}
            elif key in walls_in_new:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                walls_in_new[key]["Area"] += wall_area
                walls_in_new[key]["Volume"] += wall_volume

        if wall_duplicationTypeMark == "Existing":
            key = "Existing/קיים"
            if key not in walls_in_new:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                walls_in_new[key] = {"Area": wall_area, "Volume": wall_volume}
            elif key in walls_in_new:
                wall_area = area_param.AsDouble() * 0.092903
                wall_volume = volume_param.AsDouble() * 0.0283168466
                walls_in_new[key]["Area"] += wall_area
                walls_in_new[key]["Volume"] += wall_volume

        if wall_duplicationTypeMark == "Concrete":
            if wall_type_comments == "FrIN":
                wall_key = "קירות פנימיים מבטון/Walls-In"
                if wall_key not in walls_in_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_in_new[wall_key] = {"Area": wall_area, "Volume": wall_volume}
                elif wall_key in walls_in_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_in_new[wall_key]["Area"] += wall_area
                    walls_in_new[wall_key]["Volume"] += wall_volume

        if wall_duplicationTypeMark == "Concrete":
            if wall_type_comments == "FrOut":
                wall_key = "קירות חוץ מבטון/Walls-Out"
                if wall_key not in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key] = {"Area": wall_area, "Volume": wall_volume}
                elif wall_key in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key]["Area"] += wall_area
                    walls_out_new[wall_key]["Volume"] += wall_volume

        if wall_duplicationTypeMark in ["Concrete-P", "Concrete-WR"]:
            if wall_type_comments == "FrIN":
                wall_key = "קירות פנימיים מבטון/Walls-In"
                if wall_key not in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key] = {"Area": wall_area, "Volume": wall_volume}
                elif wall_key in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key]["Area"] += wall_area
                    walls_out_new[wall_key]["Volume"] += wall_volume


        if wall_duplicationTypeMark in ["Concrete-P", "Concrete-WR"]:
            if wall_type_comments == "FrOut":
                wall_key = "קירות חוץ מבטון/Walls-Out"
                if wall_key not in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key] = {"Area": wall_area, "Volume": wall_volume}
                elif wall_key in walls_out_new:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    walls_out_new[wall_key]["Area"] += wall_area
                    walls_out_new[wall_key]["Volume"] += wall_volume

        if wall_duplicationTypeMark == "Precast":
            if wall_type_comments == "FrOut":
                wall_key = "קירות טרומיים/Concrete-Precast"
                if wall_key not in precast_elements:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    precast_elements[wall_key] = {"Area": wall_area, "Volume": wall_volume}
                elif wall_key in precast_elements:
                    wall_area = area_param.AsDouble() * 0.092903
                    wall_volume = volume_param.AsDouble() * 0.0283168466
                    precast_elements[wall_key]["Area"] += wall_area
                    precast_elements[wall_key]["Volume"] += wall_volume



getting_Area_Volume_walls(wall_collector)

"""________Foundations________"""
foundation_collector = FilteredElementCollector(doc). \
    OfCategory(BuiltInCategory.OST_StructuralFoundation). \
    WhereElementIsNotElementType(). \
    ToElements()


def getting_Length_Volume_Count(found_list):
    for el in found_list:
        if el.Category.Name == "Structural Foundations":
            if isinstance(el, FamilyInstance):
                el_type_id = el.GetTypeId()
                type_elem = doc.GetElement(el_type_id)
                if type_elem:
                    parameter_Duplication = type_elem.LookupParameter("Duplication Type Mark").AsString()
                    if not parameter_Duplication:
                        parameter = el.LookupParameter("Length")
                        parameter_vol = el.LookupParameter("Volume")
                        parameter_Descr = type_elem.LookupParameter("Description").AsValueString()
                        key = "DTM empty Foundation"
                        if key not in Dipuns:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Found_without_DTM[key] = {'Length': parameter_value, 'Volume': parameter_value_vol,
                                                      'Count': 1}
                        else:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Found_without_DTM[key]['Length'] += parameter_value
                            Found_without_DTM[key]['Volume'] += parameter_value_vol
                            Found_without_DTM[key]['Count'] += 1

                    if parameter_Duplication == "Dipun":
                        key = "Pile Dipun "
                        parameter = el.LookupParameter("Length")
                        parameter_vol = el.LookupParameter("Volume")
                        parameter_Descr = type_elem.LookupParameter("Description").AsValueString()
                        if key+parameter_Descr not in Dipuns:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Dipuns[key+parameter_Descr] = {'Length': parameter_value, 'Volume': parameter_value_vol,
                                                       'Count': 1}
                        else:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Dipuns[key+parameter_Descr]['Length'] += parameter_value
                            Dipuns[key+parameter_Descr]['Volume'] += parameter_value_vol
                            Dipuns[key+parameter_Descr]['Count'] += 1


                    elif parameter_Duplication == "Bisus":
                        key = "Pile Bisus "
                        parameter = el.LookupParameter("Length")
                        parameter_vol = el.LookupParameter("Volume")
                        parameter_Descr = type_elem.LookupParameter("Description").AsValueString()

                        if key+parameter_Descr not in Bisus:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Bisus[key+parameter_Descr] = {'Length': parameter_value, 'Volume': parameter_value_vol,
                                                      'Count': 1}
                        else:
                            parameter_value = round(parameter.AsDouble() * 0.3048)
                            parameter_value_vol = parameter_vol.AsDouble() * 0.0283168466
                            Bisus[key+parameter_Descr]['Length'] += parameter_value
                            Bisus[key+parameter_Descr]['Volume'] += parameter_value_vol
                            Bisus[key+parameter_Descr]['Count'] += 1

                else:
                    pass
            # For Floor Types ( Rafsody, Head, BasicPlate )
            elif isinstance(el, Floor):
                el_type_id = el.GetTypeId()
                # foundation_element = doc.GetElement(el)
                foundation_type = el.FloorType
                foundation_duplicationTypeMark = foundation_type.LookupParameter("Duplication Type Mark").AsString()
                if foundation_duplicationTypeMark == "Rafsody":
                    key = "Rafsody/רפסודה"
                    if key not in Rafsody:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Rafsody[key] = {"Area": foundation_area, "Volume": foundation_volume}
                    elif key in Rafsody:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Rafsody[key]["Area"] += foundation_area
                        Rafsody[key]["Volume"] += foundation_volume

                if foundation_duplicationTypeMark == "Basic Plate":
                    key = "Basic Plate/פלטות יסוד"
                    if key not in Basic_Plate:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Basic_Plate[key] = {"Area": foundation_area,
                                            "Volume": foundation_volume}
                    elif key in Basic_Plate:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Basic_Plate[key]["Area"] += foundation_area
                        Basic_Plate[key]["Volume"] += foundation_volume
                if foundation_duplicationTypeMark == "Head":
                    key = "Foundation Head/ראשי כלונס"
                    if key not in Found_Head:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Found_Head[key] = {"Area": foundation_area, "Volume": foundation_volume}
                    elif key in Found_Head:
                        foundation_area = el.LookupParameter("Area").AsDouble() * 0.092903
                        foundation_volume = el.LookupParameter("Volume").AsDouble() * 0.0283168466
                        Found_Head[key]["Area"] += foundation_area
                        Found_Head[key]["Volume"] += foundation_volume

    return Dipuns, Bisus, Rafsody, Basic_Plate, Found_Head, Found_without_DTM


getting_Length_Volume_Count(foundation_collector)
"""________Creating_Excel_File________"""
df_found_dipuns = pd.DataFrame.from_dict(Dipuns, orient="index", columns=["Length", "Volume", "Count"])
df_found_dipuns_sorted = df_found_dipuns.copy()
# df_found_dipuns_sorted.index = pd.to_numeric(df_found_dipuns_sorted.index, errors="coerce")
df_found_dipuns_sorted = df_found_dipuns_sorted.sort_index().fillna(0)

df_found_bisus = pd.DataFrame.from_dict(Bisus, orient="index", columns=["Length", "Volume", "Count"])
df_found_bisus_sorted = df_found_bisus.copy()
# df_found_bisus_sorted.index = pd.to_numeric(df_found_bisus_sorted.index, errors="coerce")
df_found_bisus_sorted = df_found_bisus_sorted.sort_index().fillna(0)

df_found_without_DTM = pd.DataFrame.from_dict(Found_without_DTM, orient="index", columns=["Length", "Volume", "Count"])

df_found_Slurry_Bisus = pd.DataFrame.from_dict(slurry_Bisus, orient="index", columns=["Area", "Volume"])
df_found_Slurry_Dipuns = pd.DataFrame.from_dict(slurry_Dipun, orient="index", columns=["Area", "Volume"])

df_Basic_Plate = pd.DataFrame.from_dict(Basic_Plate, orient="index", columns=["Area", "Volume"])
df_Rafsody = pd.DataFrame.from_dict(Rafsody, orient="index", columns=["Area", "Volume"])
df_Found_Head = pd.DataFrame.from_dict(Found_Head, orient="index", columns=["Area", "Volume"])

df_floors_up = pd.DataFrame.from_dict(floors_up, orient="index", columns=["Area", "Volume"])
df_floors_down = pd.DataFrame.from_dict(floors_down, orient="index", columns=["Area", "Volume"])

df_beams = pd.DataFrame.from_dict(beams, orient="index", columns=["Volume", "Count"])

df_precast_elements = pd.DataFrame.from_dict(precast_elements, orient="index", columns=["Area", "Volume", "Count"])

df_edges = pd.DataFrame.from_dict(Slabedge, orient="index", columns=["Volume", "Length"])

df_stairs = pd.DataFrame.from_dict(Stairs, orient="index", columns=["Volume"])

df_columns = pd.DataFrame.from_dict(columns, orient="index", columns=["Volume", "Count"])

df_walls_in = pd.DataFrame.from_dict(walls_in_new, orient="index", columns=["Area", "Volume"])
df_walls_out = pd.DataFrame.from_dict(walls_out_new, orient="index", columns=["Area", "Volume"])

"""Inserting new nam of type_element"""
# name_up = "Floors"
df_name_floors_up = pd.DataFrame(index=["Floors/קומות"])
# name_down = "Floors_Down"
df_name_floors_down = pd.DataFrame(index=["Floors_Down"])
# name_beams = "Beams"
df_name_beams = pd.DataFrame(index=['Beams/קורות'])
# name_walls = "Walls"
df_name_walls = pd.DataFrame(index=['Walls/קירות'])
# name_foundation = "Foundation"
df_name_Dipuns = pd.DataFrame(index=['Dipuns/כלונסאות דיפון'])

df_name_Bisus = pd.DataFrame(index=['Bisus/כלונסאות ביסוס'])

df_name_Slurry_walls = pd.DataFrame(index=['Slurry walls/קירות דש'])

df_name_Columns = pd.DataFrame(index=['Columns/עמודי בטון'])

df_name_precast_elements = pd.DataFrame(index=['Precast elements/אלמנטים מראש'])

df_name_Stairs = pd.DataFrame(index=['Stairs/מדרגות'])

df_name_Slabedge = pd.DataFrame(index=['Wuta/ווטות מתחת לרצפה'])
df_name_BasicPlate = pd.DataFrame(index=['Basic_Plate/פלטות יסוד'])
df_name_Rafsody = pd.DataFrame(index=['Rafsody/רפסודה'])
# df_name_Found_Head = pd.DataFrame(index=['Foundation Head/קורות קשר'])
df_name_FOUNDATIONS = pd.DataFrame(index=['Foundations/יסודות'])

# dataframe_total = [(df_name_FOUNDATIONS, df_found_Slurry_Bisus, df_found_Slurry_Dipuns),
#                    (df_name_Dipuns, df_found_dipuns_sorted),
#                    (df_name_Bisus, df_found_bisus_sorted),
#                    (df_Rafsody, df_Basic_Plate,df_Found_Head,df_found_without_DTM),
#                    (df_name_Slabedge, df_edges),
#                    (df_name_beams, df_beams),
#                    (df_name_floors_up, df_floors_up, df_floors_down),
#                    (df_name_Columns, df_columns),
#                    (df_name_walls, df_walls_in, df_walls_out),
#                    (df_name_Stairs, df_stairs),
#                    ]
dataframe_total = [
    (df_name_FOUNDATIONS, df_found_Slurry_Bisus, df_found_Slurry_Dipuns, df_found_dipuns_sorted,
     df_found_bisus_sorted, df_Rafsody, df_Basic_Plate, df_Found_Head, df_found_without_DTM, df_edges),
    (df_name_precast_elements, df_precast_elements),
    (df_name_beams, df_beams),
    (df_name_floors_up, df_floors_up, df_floors_down),
    (df_name_Columns, df_columns),
    (df_name_walls, df_walls_in, df_walls_out),
    (df_name_Stairs, df_stairs),
    ]
non_empy_total_df = []

for tuple_ in dataframe_total:
    if len(tuple_) == 2:
        df_name, df1 = tuple_
        if not df1.empty:
            non_empy_total_df.append(pd.concat([df_name, df1]))
    elif len(tuple_) == 10:
        df_name, df1, df2, df3, df4, df5, df6, df7, df8, df9 = tuple_
        if not df1.empty or not df2.empty or not df3.empty or not df4.empty or not df5.empty or not df6.empty or not df7.empty or not df8.empty or not df9.empty:
            non_empy_total_df.append(
                pd.concat([df_name, df1, df2, df3, df4, df5, df6, df7, df8, df9]))
    elif len(tuple_) == 3:
        df_name, df1, df2 = tuple_
        if not df1.empty or not df2.empty:
            non_empy_total_df.append(pd.concat([df_name, df1, df2]))

df = pd.concat(non_empy_total_df, axis=0, sort=False).round(1)

df_to_sum_without_foundation_volume = [df_beams["Volume"],
                                       df_floors_up["Volume"],
                                       df_floors_down["Volume"],
                                       df_columns["Volume"],
                                       df_walls_in["Volume"],
                                       df_walls_out["Volume"],
                                       df_stairs["Volume"],
                                       ]

df_to_sum_without_foundation_area = [df_floors_up["Area"],
                                     df_floors_down["Area"],
                                     df_walls_in["Area"],
                                     df_walls_out["Area"]]

df_to_sum_found = [
    df_found_dipuns_sorted["Volume"],
    df_found_bisus_sorted["Volume"],
    df_found_Slurry_Bisus["Volume"],
    df_found_Slurry_Dipuns["Volume"],
    df_edges["Volume"],
    df_Basic_Plate["Volume"],
    df_Found_Head["Volume"],
    df_Rafsody["Volume"]

]

df_sum_volume = pd.concat(df_to_sum_without_foundation_volume).sum().round(1)
df_sum_area = pd.concat(df_to_sum_without_foundation_area).sum().round(1)
df_sum_volume_foundation = round(pd.concat(df_to_sum_found).sum(), 1)
# df_summed_without_foundation = pd.DataFrame({"Summed volume": [df_sum]})
# col_1_sum = df['Area'].sum()
# col_2_sum = df['Volume'].sum()

total_row = pd.Series({"Area": df_sum_area, "Volume": df_sum_volume, "Count": ""},
                      name="Total without foundation/סך הכל ללא בסיס")
total_row_found = pd.Series({"Area": "", "Volume": df_sum_volume_foundation, "Count": ""},
                            name="Total foundation/בסיס כולל")

df = df._append(total_row_found)
df = df._append(total_row)

df = df.fillna('')
# changing position of columns
df = df.reindex(columns=["Area", "Volume", "Count", "Length"])
df = df.rename(columns={"Area" : "Area/אזור", "Volume" : "Volume/כרך", "Count":"לספור/Count","Length":"Length/אורך"})

writer = pd.ExcelWriter(IN[1], engine='openpyxl')
df.to_excel(writer, sheet_name="Test", index=True)

workbook = writer.book
worksheet = writer.sheets['Test']
column_index = 1  # column A`

for col_num, column_title in enumerate(df.columns, 1):
    cell = worksheet.cell(row=1, column=col_num + 1)
    cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

for row_index in range(1, df.shape[0] + 2):
    cell = worksheet.cell(row=row_index, column=column_index)
    cell.fill = PatternFill(start_color="00FFFFCC", end_color="00FFFFCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# set width of first column to 20
column_letter = get_column_letter(column_index)
column_dimension = worksheet.column_dimensions[column_letter]
column_dimension.width = 25

red_fill = PatternFill(start_color='00CCCCFF', end_color='00CCCCFF', fill_type='solid')
_fill = PatternFill(start_color='00CCCCFF', end_color='00CCCCFF', fill_type='solid')

for row in worksheet.iter_rows():
    for cell in row:
        if cell.value in ['Foundations/יסודות', "Floors/קומות", 'Beams/קורות', 'Walls/קירות',
                            'Columns/עמודי בטון', 'Stairs/מדרגות','Precast elements/אלמנטים מראש']:
            cell.fill = red_fill
        elif cell.value in ["Total foundation/בסיס כולל", "Total without foundation/סך הכל ללא בסיס"]:
            cell.fill = _fill

writer._save()
